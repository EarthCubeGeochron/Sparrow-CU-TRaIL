# -*- coding: utf-8 -*-
"""
Created on Mon Jun  6 13:25:10 2022

@author: Peter
"""

from sparrow.import_helpers import BaseImporter
from sparrow.util import relative_path
from yaml import load
from rich import print
import pandas as pd
import openpyxl
import shutil

def add_column(ws, col, label, thin):
        ws.cell(row=2, column=col, value=label)
        ws.cell(row=2, column=col).alignment = openpyxl.styles.Alignment(horizontal='center',
                                                                         vertical='center',
                                                                         wrapText=True)
        ws.cell(row=2, column=col).font = openpyxl.styles.Font(name='Helvetica Neue',
                                                               bold=True)
        ws.cell(row=2, column=col).border = openpyxl.styles.Border(top=thin, bottom=thin)
        

class PublicationTable_exporter(BaseImporter):
    def __init__(self, app, data_dir, **kwargs):
        super().__init__(app)
        self.file = str(data_dir)+'/ExportPublicationTable/SampleIDs.csv'
        
        # Load the column specs; structure is {parameter: [value col, error col, unit str]}
        spec = relative_path(__file__, 'publication_table_info.yaml')
        with open(spec) as f:
            self.table_specs = load(f)
        
        # print(self.table_specs)
        self.create_table(data_dir)
    
    def query_datum(self, lab_id, datum_param):
        Session = self.db.model.session
        Sample = self.db.model.sample
        Analysis = self.db.model.analysis
        Datum = self.db.model.datum
        DatumType = self.db.model.datum_type
        res = (self.db.session.query(Datum)
               .join(Analysis)
               .join(Session)
               .join(Sample)
               .join(DatumType)
               .filter(Sample.lab_id == lab_id)
               .filter(DatumType.parameter == datum_param)
               .first())
        return res
    
    def query_datum_unit(self, lab_id, datum_param, datum_unit):
        Session = self.db.model.session
        Sample = self.db.model.sample
        Analysis = self.db.model.analysis
        Datum = self.db.model.datum
        DatumType = self.db.model.datum_type
        res = (self.db.session.query(Datum)
               .join(Analysis)
               .join(Session)
               .join(Sample)
               .join(DatumType)
               .filter(Sample.lab_id == lab_id)
               .filter(DatumType.parameter == datum_param)
               .filter(DatumType.unit == datum_unit)
               .first())
        return res
    
    def query_attribute(self, lab_id, attribute_param):
        Session = self.db.model.session
        Sample = self.db.model.sample
        Analysis = self.db.model.analysis
        Attribute = self.db.model.attribute
        res = (self.db.session.query(Attribute)
               .join(Analysis.attribute_collection)
               .join(Session)
               .join(Sample)
               .filter(Sample.lab_id == lab_id)
               .filter(Attribute.parameter == attribute_param)
               .first())
        return res
    
    def create_table(self, data_dir):
        # get samples from input file
        samples = list(pd.read_csv(self.file, header = None)[0])
        sample_dict = {}
        for sample in samples:
            try:
                aliquot = (self.db.session
                           .query(self.db.model.sample)
                           .filter_by(lab_id=sample)
                           .first().name)
            except AttributeError:
                print('No sample with lab ID:', sample)
                continue
            # split by sample and aliquot names for table formatting
            aliquot_split = aliquot.rsplit('_', 1)
            if aliquot_split[0] not in sample_dict:
                try:
                    sample_dict[aliquot_split[0]] = [(aliquot_split[1], sample)]
                # If no underscore in sample name, 
                except IndexError:
                    sample_dict.setdefault('Unknown parent sample', []).append((aliquot_split[0], sample))
            else:
                sample_dict[aliquot_split[0]].append((aliquot_split[1], sample))
        
        # Make workbook
        file_path = str(data_dir)+'/ExportPublicationTable/test.xlsx'
        temp_path = '/tmp/test.xslx'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value='Table . Publication table')
        
        # Set up column headers
        thin = openpyxl.styles.Side(border_style="thin", color="000000")
        add_column(ws, 1, 'Sample Name and Aliquot', thin)
        n = 1
        for dict_ in self.table_specs[:-1]:
            key = next(iter(dict_))
            n+=1
            add_column(ws, n, dict_[key]['column'], thin)
            ws.column_dimensions[openpyxl.utils.get_column_letter(n)].width = dict_[key]['width']
            if dict_[key]['error']:
                n+=1
                ws.column_dimensions[openpyxl.utils.get_column_letter(n)].width = dict_[key]['error_width']
                if 'error_name' in dict_[key]:
                    add_column(ws, n, dict_[key]['error_name'], thin)
                else:
                    add_column(ws, n, '± '+str(dict_[key]['error_footnote']), thin)
                
        
        # Add data one sample at a time
        row = 2
        for sample in sample_dict:
            row += 1
            ws.cell(row=row, column=1, value=sample).font = openpyxl.styles.Font(underline='single', size = '12')
            for aliquot in sample_dict[sample]:
                row += 1
                ws.cell(row=row, column=1, value=aliquot[0]).font = openpyxl.styles.Font(size = '12')
                col = 1
                for dict_ in self.table_specs[:-1]:
                    col += 1
                    key = next(iter(dict_))
                    # Default to datum
                    item = self.query_datum(aliquot[1], key)
                    # If None found, key should lead to attribute
                    if not item:
                        item = self.query_attribute(aliquot[1], key)
                    # Overwrite with correct unit if necessary
                    if 'unit' in dict_[key]:
                        item = self.query_datum_unit(aliquot[1], key, dict_[key]['unit'])
                    try:
                        ws.cell(row=row, column=col, value=item.value).font = openpyxl.styles.Font(size = '12')
                        ws.cell(row=row, column=col).alignment = openpyxl.styles.Alignment(horizontal='center')
                        if dict_[key]['error']:
                            col+=1
                            ws.cell(row=row, column=col, value=item.error).font = openpyxl.styles.Font(size = '12')
                            ws.cell(row=row, column=col).alignment = openpyxl.styles.Alignment(horizontal='center')
                            pass
                    except:
                        if dict_[key]['error']:
                            col+=1
        
        # Add gap row before footnotes
        row += 1
        for n in range(1,n+1):
            ws.cell(row=row, column=n).border = openpyxl.styles.Border(top=thin)
        
        # Add footnotes
        for note in self.table_specs[-1]['Footnotes']:
            row+=1
            ws.cell(row=row, column=1, value=note)
        
        # Save workbook
        wb.save(temp_path)
        shutil.move(temp_path, file_path)
        
        print('Table generated')