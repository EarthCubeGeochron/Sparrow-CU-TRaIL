# -*- coding: utf-8 -*-
"""
Created on Mon Jun  6 13:25:10 2022

@author: Peter
"""

from sparrow.import_helpers import BaseImporter
from sparrow.util import relative_path
from yaml import load
from rich import print
import pandas as pd
import openpyxl
import shutil

def add_column(ws, col, label, thin):
        ws.cell(row=2, column=col, value=label)
        ws.cell(row=2, column=col).alignment = openpyxl.styles.Alignment(horizontal='center',
                                                                         vertical='center',
                                                                         wrapText=True)
        ws.cell(row=2, column=col).font = openpyxl.styles.Font(name='Helvetica Neue',
                                                               bold=True)
        ws.cell(row=2, column=col).border = openpyxl.styles.Border(top=thin, bottom=thin)
        

class PublicationTableExporter(BaseImporter):
    def __init__(self, app, data_dir, file_out, **kwargs):
        super().__init__(app)
        self.file = str(data_dir)+'/ExportPublicationTable/Sample_IDs.xlsx'
        
        # Get file name from UI or input function if run from CLI
        if not file_out:
            try:
                self.file_out = input('Enter name of publication table to save: ')
            except:
                print('Please enter a file name to save.')
                return
        else:
            self.file_out = file_out
        if '.' in self.file_out:
            if not self.file_out.endswith('.xlsx'):
                self.file_out = self.file_out.split('.')[0]+'.xlsx'
        else:
            self.file_out = self.file_out +'.xlsx'
        
        print('Saving table as:', self.file_out)
        
        # Load the column specs; structure is {parameter: [value col, error col, unit str]}
        spec = relative_path(__file__, 'publication_table_info.yaml')
        with open(spec) as f:
            self.table_specs_new = load(f)
        spec = relative_path(__file__, 'publication_table_info_archive.yaml')
        with open(spec) as fi:
            self.table_specs_archive = load(fi)
        
        self.create_table(data_dir)
    
    def query_archive(self, lab_id):
        Sample = self.db.model.sample
        res = (self.db.session.query(Sample)
               .filter(Sample.lab_id == lab_id)
               .first())
        return res.from_archive
    
    def query_datum(self, lab_id, datum_param):
        Session = self.db.model.session
        Sample = self.db.model.sample
        Analysis = self.db.model.analysis
        Datum = self.db.model.datum
        DatumType = self.db.model.datum_type
        res = (self.db.session.query(Datum)
               .join(Analysis)
               .join(Session)
               .join(Sample)
               .join(DatumType)
               .filter(Sample.lab_id == lab_id)
               .filter(DatumType.parameter == datum_param)
               .first())
        return res
    
    def query_datum_unit(self, lab_id, datum_param, datum_unit):
        Session = self.db.model.session
        Sample = self.db.model.sample
        Analysis = self.db.model.analysis
        Datum = self.db.model.datum
        DatumType = self.db.model.datum_type
        res = (self.db.session.query(Datum)
               .join(Analysis)
               .join(Session)
               .join(Sample)
               .join(DatumType)
               .filter(Sample.lab_id == lab_id)
               .filter(DatumType.parameter == datum_param)
               .filter(DatumType.unit == datum_unit)
               .first())
        return res
    
    def query_attribute(self, lab_id, attribute_param):
        Session = self.db.model.session
        Sample = self.db.model.sample
        Analysis = self.db.model.analysis
        Attribute = self.db.model.attribute
        res = (self.db.session.query(Attribute)
               .join(Analysis.attribute_collection)
               .join(Session)
               .join(Sample)
               .filter(Sample.lab_id == lab_id)
               .filter(Attribute.parameter == attribute_param)
               .first())
        return res
    
    def create_table(self, data_dir):
        # get samples from input file
        df = pd.read_excel(self.file, header = None)
        if 0 in df.columns:
            samples = list(df[0])
        else:
            samples = []
        sample_dict = {}
        for sample in samples:
            try:
                aliquot = (self.db.session
                           .query(self.db.model.sample)
                           .filter_by(lab_id=sample)
                           .first().name)
            except AttributeError:
                print('No sample with lab ID:', sample)
                continue
            # split by sample and aliquot names for table formatting
            aliquot_split = aliquot.rsplit('_', 1)
            if aliquot_split[0] not in sample_dict:
                try:
                    sample_dict[aliquot_split[0]] = [(aliquot_split[1], sample)]
                # If no underscore in sample name, 
                except IndexError:
                    sample_dict.setdefault('Unknown parent sample', []).append((aliquot_split[0], sample))
            else:
                sample_dict[aliquot_split[0]].append((aliquot_split[1], sample))
        
        # Make workbook
        file_path = str(data_dir)+'/ExportPublicationTable//'+self.file_out
        temp_path = '/tmp//'+self.file_out
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value='Table . Publication table')
        
        # Set up column headers
        thin = openpyxl.styles.Side(border_style="thin", color="000000")
        add_column(ws, 1, 'Sample Name and Aliquot', thin)
        n = 1
        for dict_ in self.table_specs_new[:-1]:
            key = next(iter(dict_))
            n+=1
            add_column(ws, n, dict_[key]['column'], thin)
            ws.column_dimensions[openpyxl.utils.get_column_letter(n)].width = dict_[key]['width']
            if dict_[key]['error']:
                n+=1
                ws.column_dimensions[openpyxl.utils.get_column_letter(n)].width = dict_[key]['error_width']
                if 'error_name' in dict_[key]:
                    add_column(ws, n, dict_[key]['error_name'], thin)
                else:
                    add_column(ws, n, '± '+str(dict_[key]['error_footnote']), thin)
                if 'secondary_error' in dict_[key]:
                    n+=1
                    add_column(ws, n, dict_[key]['secondary_error_name'], thin)
                    ws.column_dimensions[openpyxl.utils.get_column_letter(n)].width = dict_[key]['secondary_error_width']
        
        # Add data one sample at a time
        row = 2
        for sample in sample_dict:
            row += 1
            ws.cell(row=row, column=1, value=sample).font = openpyxl.styles.Font(underline='single', size = '12')
            for aliquot in sample_dict[sample]:
                row += 1
                ws.cell(row=row, column=1, value=aliquot[0]).font = openpyxl.styles.Font(size = '12')
                col = 1
                archive = self.query_archive(aliquot[1])
                if archive:
                    self.table_specs = self.table_specs_archive
                else:
                    self.table_specs = self.table_specs_new
                for dict_ in self.table_specs[:-1]:
                    col += 1
                    key = next(iter(dict_))
                    # Default to datum
                    item = self.query_datum(aliquot[1], key)
                    # fix special case of some missing iterated dates in archived data
                    if archive and "Date" in key:
                        item = self.query_datum(aliquot[1], key)
                        if not item:
                            item = self.query_datum(aliquot[1], key.replace("Iterated", "Linear (M&D)"))
                    # If None found, key should lead to attribute
                    if not item:
                        item = self.query_attribute(aliquot[1], key)
                    # Overwrite with correct unit if necessary
                    if 'unit' in dict_[key]:
                        item = self.query_datum_unit(aliquot[1], key, dict_[key]['unit'])
                    try:
                        if 'round' in dict_[key]:
                            rounded_num = round(item.value, dict_[key]['round']) if dict_[key]['round'] != 0 else int(item.value)
                            ws.cell(row=row, column=col, value=rounded_num).font = openpyxl.styles.Font(size = '12')
                            ws.cell(row=row, column=col).alignment = openpyxl.styles.Alignment(horizontal='center')
                        else:
                            ws.cell(row=row, column=col, value=item.value).font = openpyxl.styles.Font(size = '12')
                            ws.cell(row=row, column=col).alignment = openpyxl.styles.Alignment(horizontal='center')
                        if dict_[key]['error']:
                            col+=1
                            rounded_err = round(item.error, dict_[key]['error_round']) if dict_[key]['error_round'] != 0 else int(item.error)
                            ws.cell(row=row, column=col, value=rounded_err).font = openpyxl.styles.Font(size = '12')
                            ws.cell(row=row, column=col).alignment = openpyxl.styles.Alignment(horizontal='center')
                            if 'secondary_error' in dict_[key]:
                                col+=1
                                second_err_item = self.query_datum(aliquot[1], dict_[key]['secondary_error'])
                                rounded_second_err = round(second_err_item.error, dict_[key]['secondary_error_round']) if dict_[key]['secondary_error_round'] != 0 else int(second_err_item.error)
                                ws.cell(row=row, column=col, value=rounded_second_err).font = openpyxl.styles.Font(size = '12')
                                ws.cell(row=row, column=col).alignment = openpyxl.styles.Alignment(horizontal='center')
                    except:
                        if dict_[key]['error']:
                            col+=1
        
        # Add gap row before footnotes
        row += 1
        for n in range(1,n+1):
            ws.cell(row=row, column=n).border = openpyxl.styles.Border(top=thin)
        
        # Add footnotes
        for note in self.table_specs_new[-1]['Footnotes']:
            row+=1
            ws.cell(row=row, column=1, value=note)
        
        # Save workbook
        wb.save(temp_path)
        shutil.move(temp_path, file_path)
        
        print('Table generated')